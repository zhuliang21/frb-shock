#!/usr/bin/env python3
"""
Build the current-vs-last-year table from shock_data.json and export to Excel.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SUMMARY_PATH = PROJECT_ROOT / "data" / "intermediate" / "shock_data.json"
SPEC_PATH = PROJECT_ROOT / "config" / "table_specs" / "table_vs_lastyear.json"


def load_summary() -> Dict[str, Dict[str, Any]]:
    return json.loads(SUMMARY_PATH.read_text())


def load_spec() -> Dict[str, Any]:
    return json.loads(SPEC_PATH.read_text())


def base_context(entry: Dict[str, Any]) -> Dict[str, Any]:
    ctx: Dict[str, Any] = {}
    shock = entry.get("shock_value")
    extreme = entry.get("extreme_value")

    if isinstance(shock, dict):
        values = list(shock.values())
        ctx["high"] = max(values)
        ctx["low"] = min(values)
    elif shock is not None:
        ctx["shock"] = shock
        ctx["delta"] = shock

    if isinstance(extreme, dict):
        ctx["extreme_high"] = max(extreme.values())
        ctx["extreme_low"] = min(extreme.values())
    elif extreme is not None:
        ctx["extreme"] = extreme

    return ctx


def render_value(entry: Dict[str, Any], spec: Dict[str, Any]) -> str:
    ctx = base_context(entry)

    if "delta" in ctx and spec.get("delta_scale"):
        ctx["delta"] = ctx["delta"] * spec["delta_scale"]

    template = spec["template"]
    return template.format(**ctx)


def build_table() -> Dict[str, Any]:
    summary = load_summary()
    spec = load_spec()
    columns = spec["columns"]
    factor_order: List[str] = spec["order"]

    result: Dict[str, Any] = {}

    # Find the "Value" column spec
    value_column = next(col for col in columns if col.get("header") == "Value")

    for factor in factor_order:
        entry = summary[factor]
        # Copy the original data
        result[factor] = entry.copy()
        
        # Find the display template for this factor
        value_spec = next(v for v in value_column["values"] if v["source"] == factor)
        # Add the display field
        result[factor]["display"] = render_value(entry, value_spec)

    return result


def export_to_excel(current_data: Dict[str, Any], spec: Dict[str, Any]) -> None:
    """Export comparison table to Excel with styling."""
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not installed. Skipping Excel export.")
        print("Install with: pip install openpyxl")
        return
    
    # Load history data
    history_path = PROJECT_ROOT / "data" / "history" / "table_vs_lastyear.json"
    if not history_path.exists():
        print(f"Warning: History file not found at {history_path}")
        return
    
    history_data = json.loads(history_path.read_text())
    
    # Get scenario names
    current_scenario = list(current_data.keys())[0]
    history_scenarios = sorted(history_data.keys())
    
    # Get factor order
    factor_order: List[str] = spec["order"]
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    
    # Define styles
    header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    header_font = Font(name="Inter", color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_font = Font(name="Inter", size=11)
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    
    # Build headers with line breaks
    def format_scenario_name(name: str) -> str:
        """Add line break between year and scenario type."""
        # Example: "CCAR 2024 (Severely Adverse)" -> "CCAR 2024\n(Severely Adverse)"
        import re
        match = re.match(r'(CCAR \d{4})\s+(\(.+\))', name)
        if match:
            return f"{match.group(1)}\n{match.group(2)}"
        return name
    
    headers = ["Factor"] + [format_scenario_name(s) for s in history_scenarios] + [format_scenario_name(current_scenario)]
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data rows
    for row_idx, factor in enumerate(factor_order, start=2):
        # Factor name
        cell = ws.cell(row=row_idx, column=1, value=factor)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # History columns
        col_idx = 2
        for scenario in history_scenarios:
            if factor in history_data[scenario]:
                display_value = history_data[scenario][factor].get("display", "")
            else:
                display_value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=display_value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border
            col_idx += 1
        
        # Current column
        if factor in current_data[current_scenario]:
            display_value = current_data[current_scenario][factor].get("display", "")
        else:
            display_value = ""
        cell = ws.cell(row=row_idx, column=col_idx, value=display_value)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Save workbook
    output_path = PROJECT_ROOT / "output" / "table_vs_lastyear.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved Excel -> {output_path}")


def main() -> None:
    spec = load_spec()
    table = build_table()
    
    # Wrap table with scenario name
    scenario_name = spec.get("scenario_name", "Unknown Scenario")
    output = {scenario_name: table}
    
    output_path = PROJECT_ROOT / spec["output_path"]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, indent=2))
    print(f"Saved JSON -> {output_path}")
    print(f"Scenario: {scenario_name}")
    
    # Export to Excel
    export_to_excel(output, spec)


if __name__ == "__main__":
    main()
