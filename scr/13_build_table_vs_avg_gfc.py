#!/usr/bin/env python3
"""
Build the avg/GFC comparison table with heatmap visualization.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SUMMARY_PATH = PROJECT_ROOT / "data" / "intermediate" / "shock_data.json"
SPEC_PATH = PROJECT_ROOT / "config" / "table_specs" / "table_vs_avg_gfc.json"


def load_summary() -> Dict[str, Dict[str, Any]]:
    return json.loads(SUMMARY_PATH.read_text())


def load_spec() -> Dict[str, Any]:
    return json.loads(SPEC_PATH.read_text())


def build_table() -> Dict[str, Any]:
    """Build current scenario data from shock_data.json."""
    summary = load_summary()
    spec = load_spec()
    
    result: Dict[str, Any] = {}
    
    # Extract values for each factor
    for group in spec["factor_groups"]:
        for factor_spec in group["factors"]:
            source = factor_spec["source"]
            if source in summary:
                shock_value = summary[source].get("shock_value")
                # Handle dict values
                if not isinstance(shock_value, dict):
                    # Convert to appropriate units based on factor
                    if "Treasury" in source or "Spread" in source:
                        # Convert to bps (percentage point * 100)
                        result[source] = shock_value * 100 if shock_value is not None else None
                    elif source == "VIX":
                        # VIX is already in points
                        result[source] = shock_value
                    else:
                        # Percentages remain as is
                        result[source] = shock_value
    
    return result


def get_color_for_abs_comparison(current_val: float, baseline_val: float, colors: Dict[str, Any]) -> PatternFill:
    """
    Compare absolute magnitudes. If current severity (abs value) is less than or equal to baseline,
    return green; otherwise red.
    """
    if current_val is None or baseline_val is None:
        return PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    if abs(current_val) >= abs(baseline_val):
        color_value = colors.get("green", "C7EA46")
    else:
        color_value = colors.get("red", "C94A34")

    return PatternFill(start_color=color_value, end_color=color_value, fill_type="solid")


def export_to_excel(current_data: Dict[str, Any], spec: Dict[str, Any]) -> None:
    """Export comparison table with heatmap to Excel."""
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not installed. Skipping Excel export.")
        return
    
    heatmap_colors = spec.get(
        "heatmap_colors",
        {
            "green": "C7EA46",
            "red": "C94A34",
        },
    )

    # Load history data
    history_path = PROJECT_ROOT / "data" / "history" / "table_vs_avg_gfc.json"
    if not history_path.exists():
        print(f"Warning: History file not found at {history_path}")
        return
    
    history_data = json.loads(history_path.read_text())
    ccar_avg_data = history_data.get("ccar_avg", {})
    gfc_data = history_data.get("gfc", {})
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Avg GFC Comparison"
    
    # Define styles
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Inter", color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    group_font = Font(name="Inter", size=10, bold=True)
    factor_font = Font(name="Inter", size=10)
    cell_font = Font(name="Inter", size=10)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center", indent=2)
    
    # Borders (vertical dashed only)
    border = Border(
        left=Side(style="dashed", color="CCCCCC"),
        right=Side(style="dashed", color="CCCCCC"),
        top=Side(style="dashed", color="CCCCCC"),
        bottom=Side(style="dashed", color="CCCCCC"),
    )
    border_left = Border(
        right=Side(style="dashed", color="CCCCCC"),
        top=Side(style="dashed", color="CCCCCC"),
        bottom=Side(style="dashed", color="CCCCCC"),
    )
    border_right = Border(
        left=Side(style="dashed", color="CCCCCC"),
        top=Side(style="dashed", color="CCCCCC"),
        bottom=Side(style="dashed", color="CCCCCC"),
    )
    
    # Write headers (now with two Factor columns)
    headers = ["", "Factor", "CCAR 2025 (SA)", "CCAR Avg.\n(2019-2025)", "GFC Shock", 
                "Relative to CCAR\nAvg", "Relative to GFC"]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        if col_idx == 1:
            cell.border = border_left
        elif col_idx == len(headers):
            cell.border = border_right
        else:
            cell.border = border
    
    # Write data rows
    row_idx = 2
    
    for group_spec in spec["factor_groups"]:
        group_name = group_spec["group"]
        factors = group_spec["factors"]
        
        # Remember the starting row for this group (for merging)
        group_start_row = row_idx
        
        # Write factors in this group
        for factor_idx, factor_spec in enumerate(factors):
            factor_name = factor_spec["name"]
            source = factor_spec["source"]
            template = factor_spec.get("template", "{value}")
            
            # Group name in first column (will be merged later)
            cell = ws.cell(row=row_idx, column=1, value=group_name if factor_idx == 0 else "")
            cell.font = group_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border_left
            
            # Factor name in second column
            cell = ws.cell(row=row_idx, column=2, value=factor_name)
            cell.font = factor_font
            cell.alignment = left_alignment
            cell.border = border
            
            # Current value
            current_val = current_data.get(source)
            if current_val is not None:
                display = template.format(value=current_val)
            else:
                display = ""
            cell = ws.cell(row=row_idx, column=3, value=display)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # CCAR Avg value
            avg_val = ccar_avg_data.get(source)
            if avg_val is not None:
                display = template.format(value=avg_val)
            else:
                display = ""
            cell = ws.cell(row=row_idx, column=4, value=display)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # GFC value
            gfc_val = gfc_data.get(source)
            if gfc_val is not None:
                display = template.format(value=gfc_val)
            else:
                display = ""
            cell = ws.cell(row=row_idx, column=5, value=display)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Relative to CCAR Avg (with color)
            if current_val is not None and avg_val is not None:
                # Empty cell with color only
                cell = ws.cell(row=row_idx, column=6, value="")
                cell.fill = get_color_for_abs_comparison(current_val, avg_val, heatmap_colors)
                cell.border = border
            else:
                cell = ws.cell(row=row_idx, column=6, value="")
                cell.border = border
            
            # Relative to GFC (with color)
            if current_val is not None and gfc_val is not None:
                cell = ws.cell(row=row_idx, column=7, value="")
                cell.fill = get_color_for_abs_comparison(current_val, gfc_val, heatmap_colors)
                cell.border = border_right
            else:
                cell = ws.cell(row=row_idx, column=7, value="")
                cell.border = border_right
            
            row_idx += 1
        
        # Merge cells for group name
        if len(factors) > 1:
            ws.merge_cells(start_row=group_start_row, start_column=1, 
                          end_row=row_idx - 1, end_column=1)
    
    # Set column widths
    ws.column_dimensions['A'].width = 16  # Group
    ws.column_dimensions['B'].width = 18  # Factor
    ws.column_dimensions['C'].width = 16  # CCAR 2025
    ws.column_dimensions['D'].width = 16  # CCAR Avg
    ws.column_dimensions['E'].width = 16  # GFC
    ws.column_dimensions['F'].width = 18  # Relative to Avg
    ws.column_dimensions['G'].width = 18  # Relative to GFC
    
    # Set row height for header
    ws.row_dimensions[1].height = 35
    
    # Save workbook
    output_path = PROJECT_ROOT / "output" / "table_vs_avg_gfc.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved Excel -> {output_path}")


def main() -> None:
    spec = load_spec()
    table = build_table()
    
    # Wrap table with scenario name
    scenario_name = spec.get("scenario_name", "CCAR 2025 (SA)")
    output = {scenario_name: table}
    
    output_path = PROJECT_ROOT / spec["output_path"]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, indent=2))
    print(f"Saved JSON -> {output_path}")
    print(f"Scenario: {scenario_name}")
    
    # Export to Excel
    export_to_excel(table, spec)


if __name__ == "__main__":
    main()

