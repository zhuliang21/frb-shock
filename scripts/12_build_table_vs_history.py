#!/usr/bin/env python3
"""
Build the historical comparison table from shock_data.json and export to Excel.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from paths import ScenarioPaths


def load_summary(paths: ScenarioPaths) -> Dict[str, Dict[str, Any]]:
    return json.loads(paths.shock_data_json.read_text())


def load_spec(paths: ScenarioPaths) -> Dict[str, Any]:
    spec_path = paths.table_config_dir / "table_vs_history.json"
    return json.loads(spec_path.read_text())


def build_table(paths: ScenarioPaths) -> Dict[str, Any]:
    """Build current scenario data from shock_data.json."""
    summary = load_summary(paths)
    spec = load_spec(paths)
    factor_order: List[str] = spec["factor_order"]
    
    result: Dict[str, Any] = {}
    
    for factor in factor_order:
        if factor in summary:
            shock_value = summary[factor].get("shock_value")
            if isinstance(shock_value, dict):
                result[factor] = None
            else:
                result[factor] = shock_value
        else:
            result[factor] = None
    
    return result


def export_to_excel(paths: ScenarioPaths, current_data: Dict[str, Any], spec: Dict[str, Any]) -> None:
    """Export historical comparison table to Excel with styling."""
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not installed. Skipping Excel export.")
        return
    
    history_path = paths.history_dir / "table_vs_history.json"
    if not history_path.exists():
        print(f"Warning: History file not found at {history_path}")
        return
    
    history_data = json.loads(history_path.read_text())
    
    current_scenario = spec.get("scenario_name", "CCAR 2025 FRB SA")
    current_scenario_data = {current_scenario: current_data}
    
    factor_order: List[str] = spec["factor_order"]
    columns_spec = spec["columns"]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Historical Comparison"
    
    header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    header_font = Font(name="Inter", color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    unit_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    unit_font = Font(name="Inter", color="FFFFFF", size=9)
    unit_alignment = Alignment(horizontal="center", vertical="center")
    
    cell_font = Font(name="Inter", size=10)
    cell_font_bold = Font(name="Inter", size=10, bold=True)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    scenario_font = Font(name="Inter", size=10, bold=False)
    scenario_font_bold = Font(name="Inter", size=10, bold=True)
    scenario_alignment = Alignment(horizontal="left", vertical="center")
    
    border = Border(
        left=Side(style="dashed", color="000000"),
        right=Side(style="dashed", color="000000"),
        top=None,
        bottom=None
    )
    border_left = Border(
        left=None,
        right=Side(style="dashed", color="000000"),
        top=None,
        bottom=None
    )
    border_right = Border(
        left=Side(style="dashed", color="000000"),
        right=None,
        top=None,
        bottom=None
    )
    
    headers = ["Scenario"]
    units = [""]
    
    for col_spec in columns_spec[1:]:
        headers.append(col_spec.get("header", col_spec["source"]))
        units.append(col_spec.get("unit", ""))
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        if col_idx == 1:
            cell.border = border_left
        elif col_idx == len(headers):
            cell.border = border_right
        else:
            cell.border = border
    
    for col_idx, unit in enumerate(units, start=1):
        cell = ws.cell(row=2, column=col_idx, value=unit)
        cell.fill = unit_fill
        cell.font = unit_font
        cell.alignment = unit_alignment
        if col_idx == 1:
            cell.border = border_left
        elif col_idx == len(units):
            cell.border = border_right
        else:
            cell.border = border
    
    green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    ordered_scenarios = []
    
    for scenario_name in history_data.keys():
        if "Average" not in scenario_name and "Financial Crisis" not in scenario_name:
            ordered_scenarios.append(scenario_name)
    
    if "Average FRB SA (2019-2025)" in history_data:
        ordered_scenarios.append("Average FRB SA (2019-2025)")
    
    ordered_scenarios.append(current_scenario)
    
    if "Financial Crisis Historical" in history_data:
        ordered_scenarios.append("Financial Crisis Historical")
    
    row_idx = 3
    for scenario_name in ordered_scenarios:
        is_bold_row = (
            "Average" in scenario_name or 
            scenario_name == current_scenario or 
            "Financial Crisis" in scenario_name
        )
        
        if scenario_name == current_scenario:
            scenario_data = current_scenario_data[scenario_name]
            row_fill = green_fill
        else:
            scenario_data = history_data[scenario_name]
            if "Financial Crisis" in scenario_name:
                row_fill = red_fill
            else:
                row_fill = None
        
        cell = ws.cell(row=row_idx, column=1, value=scenario_name)
        cell.font = scenario_font_bold if is_bold_row else scenario_font
        cell.alignment = scenario_alignment
        cell.border = border_left
        if row_fill:
            cell.fill = row_fill
        
        for col_idx, factor in enumerate(factor_order, start=2):
            value = scenario_data.get(factor)
            
            if value is not None:
                col_spec = next((c for c in columns_spec[1:] if c["source"] == factor), None)
                if col_spec and "template" in col_spec:
                    template = col_spec["template"]
                    try:
                        display_value = template.format(shock=value)
                    except:
                        display_value = str(value)
                else:
                    display_value = str(value)
            else:
                display_value = ""
            
            cell = ws.cell(row=row_idx, column=col_idx, value=display_value)
            cell.font = cell_font_bold if is_bold_row else cell_font
            cell.alignment = cell_alignment
            if col_idx == len(factor_order) + 1:
                cell.border = border_right
            else:
                cell.border = border
            if row_fill:
                cell.fill = row_fill
        
        row_idx += 1
    
    ws.column_dimensions['A'].width = 28
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 11
    
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    
    output_path = paths.artifacts_dir / "table_vs_history.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved Excel -> {output_path}")


def main() -> None:
    paths = ScenarioPaths()
    spec = load_spec(paths)
    table = build_table(paths)
    
    scenario_name = spec.get("scenario_name", "CCAR 2025 FRB SA")
    output = {scenario_name: table}
    
    output_path = paths.current_dir / "table_vs_history.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, indent=2))
    print(f"Saved JSON -> {output_path}")
    
    export_to_excel(paths, table, spec)


if __name__ == "__main__":
    main()
