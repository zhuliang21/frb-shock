#!/usr/bin/env python3
"""
Build the avg/GFC comparison table with heatmap visualization.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict

try:
    from openpyxl import Workbook  # type: ignore[import-untyped]
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore[import-untyped]
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from paths import ScenarioPaths


def load_summary(paths: ScenarioPaths) -> Dict[str, Dict[str, Any]]:
    return json.loads(paths.shock_data_json.read_text())


def load_spec(paths: ScenarioPaths) -> Dict[str, Any]:
    spec_path = paths.table_config_dir / "table_vs_avg_gfc.json"
    return json.loads(spec_path.read_text())


def build_table(paths: ScenarioPaths) -> Dict[str, Any]:
    """Build current scenario data from shock_data.json."""
    summary = load_summary(paths)
    spec = load_spec(paths)
    
    result: Dict[str, Any] = {}
    
    for group in spec["factor_groups"]:
        for factor_spec in group["factors"]:
            source = factor_spec["source"]
            if source in summary:
                shock_value = summary[source].get("shock_value")
                if not isinstance(shock_value, dict):
                    if "Treasury" in source or "Spread" in source:
                        result[source] = shock_value * 100 if shock_value is not None else None
                    elif source == "VIX":
                        result[source] = shock_value
                    else:
                        result[source] = shock_value
    
    return result


def get_color_for_abs_comparison(current_val: float, baseline_val: float, colors: Dict[str, Any]) -> PatternFill:
    if current_val is None or baseline_val is None:
        return PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    if abs(current_val) >= abs(baseline_val):
        color_value = colors.get("green", "C7EA46")
    else:
        color_value = colors.get("red", "C94A34")

    return PatternFill(start_color=color_value, end_color=color_value, fill_type="solid")


def export_to_excel(paths: ScenarioPaths, current_data: Dict[str, Any], spec: Dict[str, Any]) -> None:
    """Export comparison table with heatmap to Excel."""
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not installed. Skipping Excel export.")
        return
    
    heatmap_colors = spec.get("heatmap_colors", {"green": "C7EA46", "red": "C94A34"})

    history_path = paths.history_dir / "table_vs_avg_gfc.json"
    if not history_path.exists():
        print(f"Warning: History file not found at {history_path}")
        return
    
    history_data = json.loads(history_path.read_text())
    ccar_avg_data = history_data.get("ccar_avg", {})
    gfc_data = history_data.get("gfc", {})
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Avg GFC Comparison"
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Inter", color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    group_font = Font(name="Inter", size=10, bold=True)
    factor_font = Font(name="Inter", size=10)
    cell_font = Font(name="Inter", size=10)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center", indent=2)
    
    border = Border(
        left=Side(style="dashed", color="CCCCCC"),
        right=Side(style="dashed", color="CCCCCC"),
        top=Side(style="dashed", color="CCCCCC"),
        bottom=Side(style="dashed", color="CCCCCC"),
    )
    border_left = Border(
        right=Side(style="dashed", color="CCCCCC"),
        top=Side(style="dashed", color="CCCCCC"),
        bottom=Side(style="dashed", color="CCCCCC"),
    )
    border_right = Border(
        left=Side(style="dashed", color="CCCCCC"),
        top=Side(style="dashed", color="CCCCCC"),
        bottom=Side(style="dashed", color="CCCCCC"),
    )
    
    scenario_name = spec.get("scenario_name", "CCAR 2025 (SA)")
    headers = ["", "Factor", scenario_name, "CCAR Avg.\n(2019-2025)", "GFC Shock", 
                "Relative to CCAR\nAvg", "Relative to GFC"]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        if col_idx == 1:
            cell.border = border_left
        elif col_idx == len(headers):
            cell.border = border_right
        else:
            cell.border = border
    
    row_idx = 2
    
    for group_spec in spec["factor_groups"]:
        group_name = group_spec["group"]
        factors = group_spec["factors"]
        
        group_start_row = row_idx
        
        for factor_idx, factor_spec in enumerate(factors):
            factor_name = factor_spec["name"]
            source = factor_spec["source"]
            template = factor_spec.get("template", "{value}")
            
            cell = ws.cell(row=row_idx, column=1, value=group_name if factor_idx == 0 else "")
            cell.font = group_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border_left
            
            cell = ws.cell(row=row_idx, column=2, value=factor_name)
            cell.font = factor_font
            cell.alignment = left_alignment
            cell.border = border
            
            current_val = current_data.get(source)
            if current_val is not None:
                display = template.format(value=current_val)
            else:
                display = ""
            cell = ws.cell(row=row_idx, column=3, value=display)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border
            
            avg_val = ccar_avg_data.get(source)
            if avg_val is not None:
                display = template.format(value=avg_val)
            else:
                display = ""
            cell = ws.cell(row=row_idx, column=4, value=display)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border
            
            gfc_val = gfc_data.get(source)
            if gfc_val is not None:
                display = template.format(value=gfc_val)
            else:
                display = ""
            cell = ws.cell(row=row_idx, column=5, value=display)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border
            
            if current_val is not None and avg_val is not None:
                cell = ws.cell(row=row_idx, column=6, value="")
                cell.fill = get_color_for_abs_comparison(current_val, avg_val, heatmap_colors)
                cell.border = border
            else:
                cell = ws.cell(row=row_idx, column=6, value="")
                cell.border = border
            
            if current_val is not None and gfc_val is not None:
                cell = ws.cell(row=row_idx, column=7, value="")
                cell.fill = get_color_for_abs_comparison(current_val, gfc_val, heatmap_colors)
                cell.border = border_right
            else:
                cell = ws.cell(row=row_idx, column=7, value="")
                cell.border = border_right
            
            row_idx += 1
        
        if len(factors) > 1:
            ws.merge_cells(start_row=group_start_row, start_column=1, 
                          end_row=row_idx - 1, end_column=1)
    
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    
    ws.row_dimensions[1].height = 35
    
    output_path = paths.artifacts_dir / "table_vs_avg_gfc.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved Excel -> {output_path}")


def main() -> None:
    paths = ScenarioPaths()
    spec = load_spec(paths)
    table = build_table(paths)
    
    scenario_name = spec.get("scenario_name", "CCAR 2025 (SA)")
    output = {scenario_name: table}
    
    output_path = paths.current_dir / "table_vs_avg_gfc.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, indent=2))
    print(f"Saved JSON -> {output_path}")
    
    export_to_excel(paths, table, spec)


if __name__ == "__main__":
    main()
